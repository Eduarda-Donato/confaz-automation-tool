import pandas as pd
from openpyxl import load_workbook 

from excel_formatter import format_header, format_data_cell


def write_to_excel(df: pd.DataFrame, output_path: str, month_year: str, template_path: str):
    wb = load_workbook(template_path)
    ws = wb.active

    next_line = ws.max_row + 1
    ws.insert_rows(next_line)

    ws.merge_cells(f'A{next_line}:N{next_line}')
    title_cell = ws[f'A{next_line}']
    title_cell.value = f'BENEFÍCIOS - {month_year}'

    for col_ascii in range(ord('A'), ord('N') + 1):
        col = chr(col_ascii)
        cell = ws[f'{col}{next_line}']
        format_header(cell)

    start_data_row = next_line + 1

    for i, row in df.iterrows():
        for j, value in enumerate(row):
            cell = ws.cell(row=start_data_row + i, column=j + 1)
            cell.value = value
            is_date = isinstance(value, pd.Timestamp)
            format_data_cell(cell, is_date=is_date)

    wb.save(output_path)